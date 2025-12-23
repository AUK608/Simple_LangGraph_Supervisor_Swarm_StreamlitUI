import datetime
from collections import defaultdict
from typing import Callable

from langchain_core.runnables import RunnableConfig
from langgraph.checkpoint.memory import InMemorySaver
from langgraph.prebuilt import create_react_agent
from langgraph_swarm import create_handoff_tool, create_swarm
from langchain_groq import ChatGroq
import torch
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import os
import traceback
from dotenv import load_dotenv

load_dotenv()

# API Key Configuration 
groq_api = os.getenv('GROQ_API_KEY')

torch.classes.__path__ = [os.path.join(torch.__path__[0], 'torch', '_classes.py')]

#model = ChatOpenAI(model="gpt-4o")
model = ChatGroq(
    model="llama-3.3-70b-versatile",
    #model="llama3-70b-8192",
    temperature=0
)

# Mock data for tools
RESERVATIONS = defaultdict(lambda: {"flight_info": {}, "hotel_info": {}})

df_flight = pd.read_excel(r"C:\Tasks\Multimodal_Agentic_RAG\LangGraph_Supervisor_Swarm\Inputs\Sample_LG_Swarm_Data.xlsx",sheet_name="Flight_Info")
df_hotel = pd.read_excel(r"C:\Tasks\Multimodal_Agentic_RAG\LangGraph_Supervisor_Swarm\Inputs\Sample_LG_Swarm_Data.xlsx",sheet_name="Hotel_Info")

FLIGHTS = []
HOTELS = []

for i in range(len(df_flight)):
    df_f_d = {}
    df_f_d['departure_airport'] = df_flight['departure_airport'].iloc[i]
    df_f_d['arrival_airport'] = df_flight['arrival_airport'].iloc[i]
    df_f_d['airline'] = df_flight['airline'].iloc[i]
    df_f_d['date'] = df_flight['date'].iloc[i].date().isoformat()
    df_f_d['id'] = str(df_flight['id'].iloc[i])
    FLIGHTS.append(df_f_d)

for i in range(len(df_hotel)):
    df_f_h = {}
    df_f_h['location'] = df_hotel['location'].iloc[i]
    df_f_h['name'] = df_hotel['name'].iloc[i]
    df_f_h['neighborhood'] = df_hotel['neighborhood'].iloc[i]
    df_f_h['date'] = df_hotel['date'].iloc[i].date().isoformat()
    df_f_h['id'] = str(df_hotel['id'].iloc[i])
    HOTELS.append(df_f_h)

#print("FLIGHTS = = = = ", FLIGHTS)
#print("HOTELS = = = = ", HOTELS)

def save_reservations(info_dict, sheet_name):
    """
    function to save the reservations in excel file
    """
    try:
        excel_file = "C:\\Tasks\\Multimodal_Agentic_RAG\\LangGraph_Supervisor_Swarm\\Outputs\\LangGraph_Customer_Support_Output.xlsx"
        df = pd.DataFrame(info_dict)

        if os.path.isfile(excel_file):
            wb = openpyxl.load_workbook(excel_file)
            
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]  # declare the active sheet 
                # append the dataframe results to the current excel file
                for row in dataframe_to_rows(df, header = False, index = False):
                    sheet.append(row)
                wb.save(excel_file)  # save workbook
                wb.close()
            else:
                with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
                    # Write the new DataFrame to a new sheet
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(path = excel_file, engine = 'openpyxl') as writer:
                df.to_excel(writer, index = False, sheet_name = sheet_name)
        
        return "Successful"
    except Exception as e:
        print("Error in Saving Excel = = = = ", str(traceback.format_exc()))
        return "Unsuccessful"


# Flight tools
def search_flights(
    departure_airport: str,
    arrival_airport: str,
    date: str,
) -> list[dict]:
    """Search flights.

    Args:
        departure_airport: 3-letter airport code for the departure airport. If unsure, use the biggest airport in the area
        arrival_airport: 3-letter airport code for the arrival airport. If unsure, use the biggest airport in the area
        date: YYYY-MM-DD date
    """
    # return all flights for simplicity
    print("--departure_airport = = = = ", departure_airport)
    print("--arrival_airport = = = = ", arrival_airport)
    print("--date = = = = ", date)
    return FLIGHTS


def book_flight(
    flight_id: str,
    config: RunnableConfig,
):
    """Book a flight."""
    print("--flight_id = = = ", flight_id)
    user_id = config["configurable"].get("user_id")
    print("--user_id = = = ", user_id)
    flight = [flight for flight in FLIGHTS if str(flight["id"]) == str(flight_id)][0]
    print("--flight = = = ", flight)
    RESERVATIONS[user_id]["flight_info"] = flight
    print("--Reservations = = = ", RESERVATIONS)

    userdict = flight
    userlist = list(userdict.items())
    userlist.insert(0,('user_id',str(user_id)))
    userdict = dict(userlist)
    userdict = dict([(str(k),[str(v)]) for k, v in userdict.items()])

    res = save_reservations(userdict, "Flight_Reservations")
    print("--res = = = ", res)

    if res == "Successful":
        return "Successfully Booked the Flight and Details are : " + str(userdict)
    else:
        return "Flight Booking was Unsuccessful"


# Hotel tools
def search_hotels(location: str) -> list[dict]:
    """Search hotels.

    Args:
        location: offical, legal city name (proper noun)
    """
    # return all hotels for simplicity
    print("++location = = = = ", location)
    return HOTELS


def book_hotel(
    hotel_id: str,
    config: RunnableConfig,
):
    """Book a hotel"""
    print("++hotel_id = = = ", hotel_id)
    user_id = config["configurable"].get("user_id")
    print("++user_id = = = ", user_id)
    hotel = [hotel for hotel in HOTELS if str(hotel["id"]) == str(hotel_id)][0]
    print("++hotel = = = ", hotel)
    RESERVATIONS[user_id]["hotel_info"] = hotel
    print("++Reservations = = = ", RESERVATIONS)

    userdict = hotel
    userlist = list(userdict.items())
    userlist.insert(0,('user_id',str(user_id)))
    userdict = dict(userlist)
    userdict = dict([(str(k),[str(v)]) for k, v in userdict.items()])

    res = save_reservations(userdict, "Hotel_Reservations")
    print("++res = = = ", res)

    if res == "Successful":
        return "Successfully Booked the Hotel and Details are : " + str(userdict)
    else:
        return "Hotel Booking was Unsuccessful"

def get_reservations(config: RunnableConfig):
    """returns current reservations"""
    excel_file = "C:\\Tasks\\Multimodal_Agentic_RAG\\LangGraph_Supervisor_Swarm\\Outputs\\LangGraph_Customer_Support_Output.xlsx"
    wb = openpyxl.load_workbook(excel_file)
    user_id = config["configurable"].get("user_id")
    print("==user_id = = = ", user_id)
    flight_info = ""
    hotel_info = ""
    
    if "Flight_Reservations" in wb.sheetnames:
        df_f = pd.read_excel(excel_file,sheet_name="Flight_Reservations")
        df_f_u = df_f[df_f['user_id'].astype(str)==str(user_id)]
        print("--df_f_u = = = ", df_f_u)

        if df_f_u.empty:
            flight_info = "No Flight Reservations Found for Current User ID : " + str(user_id)
        else:
            df_f_u = df_f_u.to_dict('records')
            flight_info = "Flight Reservations Found for Current User ID as : " + str(df_f_u)
    
    if "Hotel_Reservations" in wb.sheetnames:
        df_h = pd.read_excel(excel_file,sheet_name="Hotel_Reservations")
        df_h_u = df_h[df_h['user_id'].astype(str)==str(user_id)]
        print("++df_h_u = = = ", df_h_u)

        if df_h_u.empty:
            hotel_info = "No Hotel Reservations Found for Current User ID : " + str(user_id)
        else:
            df_h_u = df_h_u.to_dict('records')
            hotel_info = "Hotel Reservations Found for Current User ID as : " + str(df_h_u)

    print("==flight and hotel info = = = = = ", str(flight_info) + "\n" + str(hotel_info))
    return str(flight_info) + "\n" + str(hotel_info)



# Define handoff tools
transfer_to_hotel_assistant = create_handoff_tool(
    agent_name="hotel_assistant",
    description="Transfer user to the hotel-booking assistant that can show booked history, search for and book hotels.",
)
transfer_to_flight_assistant = create_handoff_tool(
    agent_name="flight_assistant",
    description="Transfer user to the flight-booking assistant that can show booked history, search for and book flights.",
)


# Define agent prompt
def make_prompt(base_system_prompt: str) -> Callable[[dict, RunnableConfig], list]:
    def prompt(state: dict, config: RunnableConfig) -> list:
        user_id = config["configurable"].get("user_id")
        #print("--RESERVATIONS = = = ", RESERVATIONS)
        current_reservation = RESERVATIONS[user_id]
        #print("--current_reservation = = = ", current_reservation)
        system_prompt = (
            base_system_prompt
            + f"\n\nUser's active reservation: {current_reservation}"
            + f"Today is: {datetime.datetime.now()}"
        )
        return [{"role": "system", "content": system_prompt}] + state["messages"]

    return prompt


# Define agents
flight_assistant = create_react_agent(
    model,
    [search_flights, book_flight, get_reservations, transfer_to_hotel_assistant],
    prompt=make_prompt("You are a flight searching & booking assistant which can also show booked history"),
    name="flight_assistant",
)

hotel_assistant = create_react_agent(
    model,
    [search_hotels, book_hotel, get_reservations, transfer_to_flight_assistant],
    prompt=make_prompt("You are a hotel searching & booking assistant which can also show booked history"),
    name="hotel_assistant",
)

# Compile and run!
checkpointer = InMemorySaver()
builder = create_swarm([flight_assistant, hotel_assistant], default_active_agent="flight_assistant")

# Important: compile the swarm with a checkpointer to remember
# previous interactions and last active agent
app_swarm = builder.compile(checkpointer=checkpointer, name="Swarm_Agent")
config = {"configurable": {"thread_id": "1", "user_id": "1"}}#, "recursion_limit": 100}


def invoke_swarm(prompt):
    result = app_swarm.invoke({
        "messages": [
            {
                "role": "user",
                "content": f"{prompt}"
            }
        ],
    }, config)

    for m in result["messages"]:
        m.pretty_print()

    return result["messages"][-1].content

#print("+++RESERVATIONS = = = ", RESERVATIONS)

import streamlit as st

st.title("LangGraph Agentic-SWARM")
st.markdown(
    "Get Information about Flights and Hotels with Booking Feature..."
)

if "messages" not in st.session_state:
    st.session_state.messages = []

for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])

if prompt := st.chat_input("Ask your question about Flight & Hotel Bookings..."):
    st.session_state.messages.append({"role": "user", "content": prompt})
    with st.chat_message("user"):
        st.markdown(prompt)

    with st.chat_message("assistant"):
        response = invoke_swarm(prompt)
        st.write(response)
    st.session_state.messages.append({"role": "assistant", "content": response})

